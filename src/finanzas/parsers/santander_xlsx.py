"""Parser de resúmenes xlsx descargados del homebanking de Santander Río.

Layout (relativo a las filas que vi):
  R2: 'Movimientos del resumen'
  R3: 'Tarjeta Visa Crédito terminada' + ... (no usable)
  R5-R6: 'Fecha de cierre / Fecha de vencimiento'
  R7-R8: 'Total a pagar' (ARS, USD)
  R13: 'Tarjetas incluidas en el resumen | Tarjeta de | Total en pesos | Total en dólares'
  R14+: por cada tarjeta (last4 + holder + totales)
  R24: 'Pago de tarjeta y devoluciones' → header
  R25: 'Fecha | Descripción | Cuotas | Comprobante | Monto en pesos | Monto en dólares'
  R26+: payments / promos (sin comprobante)
  R32: 'Tarjeta de {holder}' (header de la sección de la primera tarjeta)
  R33: header de columnas idéntico al de pagos
  R34+: transacciones
  ... (puede haber otra "Tarjeta de {holder2}" header y otra tabla)

La columna Fecha puede venir vacía (= heredar de la fila anterior).
Las columnas son: A=Fecha, B=Descripción, C=Cuotas (str "N de M"), D=Comprobante, E=Monto $, F=Monto USD
"""

from __future__ import annotations

import re
from datetime import date, datetime, timedelta
from pathlib import Path

from openpyxl import load_workbook

from finanzas.models import ParsedStatement, ParsedTransaction
from finanzas.parsers.base import file_sha256

DATE_DDMMYYYY = re.compile(r"^(\d{2})/(\d{2})/(\d{4})$")
CARD_LAST4_RE = re.compile(r"terminada\s+en\s+(\d{4})", re.IGNORECASE)
CUOTA_RE = re.compile(r"(\d+)\s*de\s*(\d+)", re.IGNORECASE)


def parse(path: Path) -> ParsedStatement:
    p = Path(path)
    wb = load_workbook(p, data_only=True, read_only=True)
    ws = wb.active

    rows = [list(r) for r in ws.iter_rows(values_only=True)]

    period_end = _find_date(rows, "Fecha de cierre")
    due_date = _find_date(rows, "Fecha de vencimiento")
    period_start = (period_end - timedelta(days=30)) if period_end else None

    cards: dict[str, str] = {}  # card_last4 → holder
    titular_card: str | None = None
    titular_holder: str | None = None
    # La sección "Tarjetas incluidas en el resumen" tiene UN row por card con holder en col B.
    # Hay otros rows que mencionan "terminada en NNNN" (headers de sección) pero sin holder en col B.
    # Estrategia: procesar todas, pero solo overwritear cards[card] si tenemos holder real.
    for row in rows:
        if not row or not row[0]:
            continue
        cell0 = str(row[0])
        m = CARD_LAST4_RE.search(cell0)
        if m:
            card = m.group(1)
            holder = str(row[1] or "").strip() if len(row) > 1 else ""
            holder = re.sub(r"^(Adicional de|Titular)\s+", "", holder, flags=re.IGNORECASE).strip()
            holder = re.sub(r"\s*\((Titular|Adicional)\)\s*$", "", holder, flags=re.IGNORECASE).strip()
            # Solo overwriteamos si tenemos un holder real (no vacío)
            if holder:
                cards[card] = holder
                if titular_card is None and "titular" in cell0.lower() + " " + (str(row[1]) if len(row) > 1 else "").lower():
                    titular_card = card
                    titular_holder = holder
            elif card not in cards:
                cards[card] = ""
    # Si no encontramos un titular explícito, el primero descubierto es el titular
    if titular_card is None and cards:
        titular_card = next(iter(cards))
        titular_holder = cards[titular_card]

    txs: list[ParsedTransaction] = []

    current_card: str | None = titular_card
    current_holder: str | None = titular_holder
    in_tx_section = False
    last_date: date | None = None

    for row in rows:
        if not row:
            continue
        c0 = row[0]
        c0_str = str(c0).strip() if c0 is not None else ""
        c1 = row[1] if len(row) > 1 else None
        c1_str = str(c1).strip() if c1 is not None else ""

        # Headers que abren una sección de transacciones para una tarjeta específica:
        #   "Tarjeta de Federico ..."                  → titular
        #   " Adicional de Barbara Camila Herlein"     → adicional (con o sin espacio inicial)
        is_card_header = (
            c0_str.startswith("Tarjeta de ")
            or c0_str.lower().startswith("adicional de ")
            or c0_str.lower().startswith("titular ")
        )
        if is_card_header:
            holder_in_section = c0_str
            for prefix in ("Tarjeta de ", "Adicional de ", "Titular "):
                if holder_in_section.lower().startswith(prefix.lower()):
                    holder_in_section = holder_in_section[len(prefix):]
                    break
            holder_norm = holder_in_section.strip()
            # match contra los cards descubiertos en R13+: comparar primera palabra o token completo
            matched = None
            for k, v in cards.items():
                v_low = v.lower()
                h_low = holder_norm.lower()
                if any(token in v_low for token in h_low.split() if len(token) > 3):
                    matched = (k, v)
                    break
            if matched:
                current_card, current_holder = matched
            in_tx_section = False
            last_date = None
            continue

        # Header de pagos
        if c0_str.startswith("Pago de tarjeta"):
            current_card = titular_card
            current_holder = titular_holder
            in_tx_section = False
            last_date = None
            continue

        # Row 25/33: fila de header de columnas
        if c0_str.lower() == "fecha":
            in_tx_section = True
            last_date = None
            continue

        if not in_tx_section:
            continue

        # Detectar fecha
        d = _to_date(c0)
        if d:
            last_date = d
        # si c0 está vacía, heredamos la última fecha vista
        if last_date is None:
            continue

        # Validaciones de fila de tx:
        descripcion = c1_str
        cuota_str = str(row[2]).strip() if len(row) > 2 and row[2] is not None else ""
        comprobante = str(row[3]).strip() if len(row) > 3 and row[3] is not None else None
        monto_pesos = row[4] if len(row) > 4 else None
        monto_usd = row[5] if len(row) > 5 else None

        if comprobante in {"-", "", "None"}:
            comprobante = None

        amount, currency = _pick_amount(monto_pesos, monto_usd)
        if amount is None:
            continue
        if not descripcion:
            continue

        inst_cur = inst_total = None
        m_q = CUOTA_RE.search(cuota_str)
        if m_q:
            inst_cur = int(m_q.group(1))
            inst_total = int(m_q.group(2))

        txs.append(
            ParsedTransaction(
                posted_at=last_date,
                description_raw=descripcion,
                comprobante=comprobante,
                amount=amount,
                currency=currency,
                card_last4=current_card or "0000",
                holder_name=current_holder,
                installment_current=inst_cur,
                installment_total=inst_total,
            )
        )

    total_ars = sum(t.amount for t in txs if t.currency == "ARS")
    total_usd = sum(t.amount for t in txs if t.currency == "USD")

    return ParsedStatement(
        bank="santander_rio",
        period_start=period_start or date(1970, 1, 1),
        period_end=period_end or date(1970, 1, 1),
        due_date=due_date,
        transactions=txs,
        source_filename=p.name,
        file_sha256=file_sha256(p),
        raw_total_ars=total_ars,
        raw_total_usd=total_usd,
    )


def _to_date(val) -> date | None:
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    if isinstance(val, str):
        s = val.strip()
        m = DATE_DDMMYYYY.match(s)
        if m:
            try:
                return date(int(m.group(3)), int(m.group(2)), int(m.group(1)))
            except ValueError:
                return None
    return None


def _find_date(rows: list[list], label: str) -> date | None:
    """Busca la primera celda que diga exactamente `label` y devuelve la fecha de la fila siguiente
    en la misma columna (caso del layout Santander R5: 'Fecha de cierre' / R6: '23/04/2026')."""
    for i, row in enumerate(rows):
        for j, c in enumerate(row):
            if c and isinstance(c, str) and c.strip() == label:
                if i + 1 < len(rows) and j < len(rows[i + 1]):
                    return _to_date(rows[i + 1][j])
    return None


def _pick_amount(monto_pesos, monto_usd) -> tuple[float | None, str]:
    if monto_usd is not None and monto_usd != 0 and monto_usd != "":
        try:
            v = _to_float(monto_usd)
            if v != 0:
                return v, "USD"
        except (TypeError, ValueError):
            pass
    if monto_pesos is not None and monto_pesos != "" and monto_pesos != 0:
        try:
            v = _to_float(monto_pesos)
            if v != 0:
                return v, "ARS"
        except (TypeError, ValueError):
            pass
    return None, "ARS"


def _to_float(v) -> float:
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    # En el xlsx los montos vienen como strings tipo "$3.108.156,97" o "U$S81,56".
    # Ojo: hay que sacar "U$S" antes que "$" para no destruir el prefijo USD.
    s = s.replace("U$S", "").replace("USD", "").replace("$", "")
    s = s.strip()
    neg = s.endswith("-") or s.startswith("-")
    s = s.lstrip("-").rstrip("-").strip()
    s = s.replace(".", "").replace(",", ".")
    if not s:
        return 0.0
    v = float(s)
    return -v if neg else v
